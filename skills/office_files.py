
# -*- coding: utf-8 -*-
"""Read xlsx and csv files: list sheets, preview rows, basic column stats.

Paths are resolved relative to the workspace directory, matching the
convention used by other skills here.
"""
import os

from talaria.providers.base import ToolSpec

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

_WORKSPACE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "workspace"))
_MAX_ROWS_PREVIEW = 50
_MAX_CHARS = 6000


def _resolve_path(path: str) -> str:
    p = str(path).strip()
    if not p:
        return ""
    if os.path.isabs(p):
        return p
    return os.path.join(_WORKSPACE_DIR, p)


def xlsx_list_sheets(path: str = "") -> str:
    """List sheet names and dimensions in an xlsx workbook."""
    if openpyxl is None:
        return "Error: openpyxl is not installed in the sandbox."
    fpath = _resolve_path(path)
    if not fpath:
        return "Error: empty path"
    if not os.path.isfile(fpath):
        return f"Error: file not found: {path}"

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        return f"Error: could not open workbook ({type(e).__name__}: {e})"

    lines = []
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append(f"- {name}: {ws.max_row} rows x {ws.max_column} cols")
    wb.close()
    return "\n".join(lines) if lines else "(no sheets found)"


def xlsx_read(path: str = "", sheet: str = "", cell_range: str = "", max_rows: int = 0) -> str:
    """Read rows from an xlsx sheet, optionally restricted to a cell range like 'A1:D20'.
    If sheet is empty, uses the first sheet. Returns a plain-text table preview."""
    if openpyxl is None:
        return "Error: openpyxl is not installed in the sandbox."
    fpath = _resolve_path(path)
    if not fpath:
        return "Error: empty path"
    if not os.path.isfile(fpath):
        return f"Error: file not found: {path}"

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        return f"Error: could not open workbook ({type(e).__name__}: {e})"

    sheet_name = str(sheet).strip()
    if sheet_name and sheet_name not in wb.sheetnames:
        wb.close()
        return f"Error: sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    try:
        limit = int(max_rows) if int(max_rows) > 0 else _MAX_ROWS_PREVIEW
    except (TypeError, ValueError):
        limit = _MAX_ROWS_PREVIEW

    rng = str(cell_range).strip()
    try:
        if rng:
            rows_iter = ws[rng]
        else:
            rows_iter = ws.iter_rows(max_row=limit)
    except Exception as e:
        wb.close()
        return f"Error: invalid range '{cell_range}' ({type(e).__name__}: {e})"

    lines = []
    count = 0
    for row in rows_iter:
        values = [("" if c.value is None else str(c.value)) for c in row]
        lines.append(" | ".join(values))
        count += 1
        if not rng and count >= limit:
            break
    wb.close()

    if not lines:
        return "(no rows)"
    result = "\n".join(lines)
    if len(result) > _MAX_CHARS:
        result = result[:_MAX_CHARS] + "\n...(truncated)"
    return result


def csv_read(path: str = "", max_rows: int = 0, delimiter: str = ",") -> str:
    """Read a csv file and return a preview: shape, column names, and first rows."""
    if pd is None:
        return "Error: pandas is not installed in the sandbox."
    fpath = _resolve_path(path)
    if not fpath:
        return "Error: empty path"
    if not os.path.isfile(fpath):
        return f"Error: file not found: {path}"

    try:
        limit = int(max_rows) if int(max_rows) > 0 else _MAX_ROWS_PREVIEW
    except (TypeError, ValueError):
        limit = _MAX_ROWS_PREVIEW

    sep = str(delimiter) or ","
    try:
        df = pd.read_csv(fpath, sep=sep, nrows=limit)
        full_shape_df = pd.read_csv(fpath, sep=sep, usecols=[0])
        total_rows = len(full_shape_df)
    except Exception as e:
        return f"Error: could not read csv ({type(e).__name__}: {e})"

    lines = [
        f"Columns ({len(df.columns)}): {', '.join(str(c) for c in df.columns)}",
        f"Rows shown: {len(df)} (file has ~{total_rows} data rows)",
        "",
        df.to_string(index=False, max_rows=limit),
    ]
    result = "\n".join(lines)
    if len(result) > _MAX_CHARS:
        result = result[:_MAX_CHARS] + "\n...(truncated)"
    return result


def csv_stats(path: str = "", delimiter: str = ",") -> str:
    """Compute basic per-column statistics for a csv file (count, dtype, min/max/mean for numeric,
    unique count for text)."""
    if pd is None:
        return "Error: pandas is not installed in the sandbox."
    fpath = _resolve_path(path)
    if not fpath:
        return "Error: empty path"
    if not os.path.isfile(fpath):
        return f"Error: file not found: {path}"

    sep = str(delimiter) or ","
    try:
        df = pd.read_csv(fpath, sep=sep)
    except Exception as e:
        return f"Error: could not read csv ({type(e).__name__}: {e})"

    lines = [f"Rows: {len(df)}, Columns: {len(df.columns)}", ""]
    for col in df.columns:
        series = df[col]
        dtype = str(series.dtype)
        n_missing = int(series.isna().sum())
        if pd.api.types.is_numeric_dtype(series):
            desc = series.describe()
            lines.append(
                f"- {col} ({dtype}, missing={n_missing}): "
                f"min={desc.get('min'):.4g} max={desc.get('max'):.4g} "
                f"mean={desc.get('mean'):.4g} std={desc.get('std'):.4g}"
            )
        else:
            n_unique = series.nunique(dropna=True)
            lines.append(f"- {col} ({dtype}, missing={n_missing}): unique={n_unique}")
    result = "\n".join(lines)
    if len(result) > _MAX_CHARS:
        result = result[:_MAX_CHARS] + "\n...(truncated)"
    return result


TOOLS: list[ToolSpec] = [
    ToolSpec(
        name="xlsx_list_sheets",
        description="List sheet names and dimensions (rows x cols) in an xlsx workbook.",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the .xlsx file, relative to the workspace."},
            },
            "required": ["path"],
        },
        handler=xlsx_list_sheets,
    ),
    ToolSpec(
        name="xlsx_read",
        description="Read rows from an xlsx sheet, optionally restricted to a cell range like 'A1:D20'. Returns a plain-text table.",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the .xlsx file, relative to the workspace."},
                "sheet": {"type": "string", "description": "Sheet name; empty uses the first sheet."},
                "cell_range": {"type": "string", "description": "Optional cell range like 'A1:D20'; empty reads from the top."},
                "max_rows": {"type": "integer", "description": "Max rows to read when no cell_range is given (default 50)."},
            },
            "required": ["path"],
        },
        handler=xlsx_read,
    ),
    ToolSpec(
        name="csv_read",
        description="Read a csv file and return a preview: column names, row count, and first rows as a table.",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the .csv file, relative to the workspace."},
                "max_rows": {"type": "integer", "description": "Max rows to preview (default 50)."},
                "delimiter": {"type": "string", "description": "Column delimiter (default ',')."},
            },
            "required": ["path"],
        },
        handler=csv_read,
    ),
    ToolSpec(
        name="csv_stats",
        description="Compute basic per-column statistics for a csv file: dtype, missing count, min/max/mean/std for numeric columns, unique count for text columns.",
        input_schema={
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the .csv file, relative to the workspace."},
                "delimiter": {"type": "string", "description": "Column delimiter (default ',')."},
            },
            "required": ["path"],
        },
        handler=csv_stats,
    ),
]
